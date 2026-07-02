"""
GSM ETL — Polymorphic Catalog Parser
======================================

Архитектура для переваривания РАЗНЫХ форматов Excel-каталогов:
  - JapaneseCatalogParser — для katalog_gsm.xlsx (JDM, 2-уровневый header)
  - PVLCatalogParser — для "Каталог подбора PVL.xlsx" (Газпромнефть, 31 колонка)
  - HeavyEquipmentParser — для будущих каталогов тяжелой техники
  - SmallEngineParser — для будущих каталогов бензопил/газонокосилок

Паттерн: Strategy — каждый парсер сам определяет, может ли он обработать файл,
и возвращает унифицированный список VehicleRecord.

Запуск:
    parser = CatalogParserFactory.get_parser(file_path)
    records = parser.parse(file_path)
    # records: List[VehicleRecord] — единый формат для всех парсеров
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# ============================================================================
# Унифицированная модель данных (output всех парсеров)
# ============================================================================

@dataclass
class FluidRecommendation:
    """Одна рекомендация масла для одного узла."""
    node_code: str                      # ENGINE, MANUAL_TRANSMISSION, etc.
    fluid_name: str                     # "G-Energy F Synth 5W-40"
    fluid_brand: Optional[str] = None   # "G-Energy"
    viscosity_sae: Optional[str] = None # "5W-40"
    recommendation_rank: int = 1        # 1=primary, 2=alt1, 3=alt2
    is_oem: bool = False
    volume_liters: Optional[str] = None # "1.9" / "0.8-1.2" / "<"
    applicability_conditions: dict = field(default_factory=dict)
    notes: Optional[str] = None


@dataclass
class VehicleRecord:
    """Унифицированная запись о ТС с рекомендациями."""
    vehicle_type: str           # passenger_car | heavy_equipment | ...
    brand: str
    model: str
    generation: Optional[str] = None
    sub_model: Optional[str] = None  # "Diesel", "Quattro"
    year_start: Optional[int] = None
    year_end: Optional[int] = None
    market: Optional[str] = None      # JDM | EU | US | RU
    attributes: dict = field(default_factory=dict)
    external_codes: dict = field(default_factory=dict)
    source: str = "import"
    recommendations: list[FluidRecommendation] = field(default_factory=list)

    def source_hash(self) -> str:
        """Стабильный хэш для дедупликации."""
        key = f"{self.vehicle_type}|{self.brand}|{self.model}|{self.generation or ''}|{self.sub_model or ''}|{self.year_start or ''}|{self.year_end or ''}"
        for code, val in sorted(self.external_codes.items()):
            key += f"|{code}:{val}"
        return hashlib.sha256(key.encode()).hexdigest()[:16]


# ============================================================================
# Базовый класс парсера (Strategy pattern)
# ============================================================================

class CatalogParser(ABC):
    """Базовый класс для всех парсеров каталогов."""

    @abstractmethod
    def detect(self, file_path: str) -> bool:
        """Возвращает True, если парсер подходит для файла."""
        ...

    @abstractmethod
    def parse(self, file_path: str) -> list[VehicleRecord]:
        """Парсит файл и возвращает список транспортных средств."""
        ...

    @staticmethod
    def _clean(v) -> Optional[str]:
        """Очистка строкового значения из Excel."""
        if v is None or pd.isna(v):
            return None
        s = str(v).strip()
        if not s or s.lower() in ('nan', 'none', '-'):
            return None
        return s

    @staticmethod
    def _normalize_brand(brand: str) -> str:
        """Нормализация названия бренда: 'HONDA (RUS) ' -> 'Honda'."""
        if not brand:
            return ""
        # Убираем (RUS), (USA), (EU) и тримим
        brand = re.sub(r'\s*\([A-Z]+\)\s*', '', brand)
        return brand.strip().title()

    @staticmethod
    def _parse_year_range(year_str: str) -> tuple[Optional[int], Optional[int]]:
        """Парсит диапазон лет: '02.10-03.12' -> (2010, 2012); '95.12-98.07' -> (1995, 1998)."""
        if not year_str:
            return None, None
        # Ищем первый паттерн XX.YY-XX.YY
        m = re.search(r'(\d{2})\.(\d{2})\s*[-–]\s*(\d{2})\.(\d{2})', year_str)
        if m:
            _, y1, _, y2 = m.groups()
            return 1900 + int(y1), 1900 + int(y2)
        # Паттерн '10-
        m = re.search(r"'?(\d{2})[-–]\s*$", year_str)
        if m:
            y = int(m.group(1))
            return 2000 + y, None
        # Паттерн '00-'05
        m = re.search(r"'?(\d{2})\s*[-–]\s*'?(\d{2})?", year_str)
        if m:
            y1, y2 = m.groups()
            return 2000 + int(y1), (2000 + int(y2) if y2 else None)
        return None, None

    @staticmethod
    def _normalize_volume(vol_str: str) -> Optional[str]:
        """Нормализация объёма: '2,1 ' -> '2.1', '4-4,5 ' -> '4-4.5'."""
        if not vol_str:
            return None
        s = str(vol_str).strip()
        # Заменяем запятую на точку
        s = s.replace(',', '.')
        # Нормализуем диапазоны
        s = re.sub(r'\s*-\s*', '-', s)
        if s in ('-', '<', '< ', '·', ''):
            return None
        return s

    @staticmethod
    def _safe_displacement(displacement: Optional[str]) -> Optional[float]:
        """Безопасно парсит объём двигателя, обрабатывая '0.65*2', '3.5/4.5', диапазоны."""
        if not displacement:
            return None
        s = str(displacement).strip().replace(',', '.')
        # Оставляем только цифры, точки, дефисы, слэши, звёздочки
        s = re.sub(r'[^\d.\-/*]', '', s)
        if not s or s in ('-', '*', '/'):
            return None
        # Если диапазон "3.5/4.5" или "4-4.5" — берём первое число
        m = re.match(r'^([\d.]+)', s)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                return None
        return None

    @staticmethod
    def _parse_viscosity(fluid_name: str) -> Optional[str]:
        """Извлекает SAE вязкость из названия масла."""
        if not fluid_name:
            return None
        m = re.search(r'(\d+W-?\d+)', fluid_name)
        return m.group(1) if m else None

    @staticmethod
    def _extract_brand_from_name(fluid_name: str) -> Optional[str]:
        """Извлекает бренд из названия: 'G-Energy F Synth 5W-40' -> 'G-Energy'."""
        if not fluid_name:
            return None
        # Известные бренды Газпромнефти
        for brand in ['G-Energy', 'Gazpromneft', 'G-Box', 'G-Truck', 'G-Force']:
            if fluid_name.startswith(brand):
                return brand
        return None


# ============================================================================
# Japanese Catalog Parser — для katalog_gsm.xlsx
# ============================================================================

class JapaneseCatalogParser(CatalogParser):
    """Парсер для японского JDM-каталога (Honda, Toyota, Nissan, ...)."""

    BRANDS = {'Honda', 'DAIHATSU', 'MAZDA', 'MITSUBISHI', 'NISSAN',
              'SUBARU', 'SUZUKI', 'TOYOTA', 'ACURA'}
    NODE_MAP = {
        'ENGINE': 'ENGINE',
        'MANUAL_TRANSMISSION': 'MANUAL_TRANSMISSION',
        'AUTO_TRANSMISSION': 'AUTO_TRANSMISSION',
        'CVT': 'CVT',
        'FRONT_DIFF': 'DIFFERENTIAL',
        'REAR_DIFF': 'DIFFERENTIAL',
        'STEERING': 'STEERING',
        'BRAKE': 'BRAKE',
        'COOLANT': 'COOLANT',
    }

    def detect(self, file_path: str) -> bool:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                if sn.strip() in self.BRANDS:
                    return True
            return False
        except Exception:
            return False

    def parse(self, file_path: str) -> list[VehicleRecord]:
        records: list[VehicleRecord] = []
        xls = pd.ExcelFile(file_path)

        for sheet_name in xls.sheet_names:
            if sheet_name.strip() not in self.BRANDS:
                continue
            brand = self._normalize_brand(sheet_name)

            # Японский формат: 2-уровневый header в rows 0-1
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            # Пропускаем 2 строки заголовков
            for i in range(2, len(df)):
                row = df.iloc[i]
                model = self._clean(row.iloc[0])
                if not model:
                    continue

                year_str = self._clean(row.iloc[1]) or ''
                year_start, year_end = self._parse_year_range(year_str)
                displacement = self._clean(row.iloc[2])
                body_number = self._clean(row.iloc[3])
                engine_code = self._clean(row.iloc[4])

                # Масла
                # cols 5-9: двигатель (объём, с фильтром, SAE, API, название масла)
                engine_volume = self._normalize_volume(self._clean(row.iloc[5]))
                engine_volume_filter = self._normalize_volume(self._clean(row.iloc[6]))
                engine_sae = self._clean(row.iloc[7])
                engine_oil = self._clean(row.iloc[9])  # фактическое название масла

                # Drive type + transmission (col 11)
                drive_trans = self._clean(row.iloc[11]) or ''

                # MT oil (cols 12-13)
                mt_volume = self._normalize_volume(self._clean(row.iloc[12]))
                mt_oil = self._clean(row.iloc[13])

                # AT oil (cols 14-15)
                at_volume = self._normalize_volume(self._clean(row.iloc[14]))
                at_oil = self._clean(row.iloc[15])

                # Front diff (cols 18-19) — безопасно, если колонок меньше
                fd_volume = self._normalize_volume(self._clean(row.iloc[18])) if len(row) > 18 else None
                fd_type = self._clean(row.iloc[19]) if len(row) > 19 else None

                # Rear diff (cols 20-21)
                rd_volume = self._normalize_volume(self._clean(row.iloc[20])) if len(row) > 20 else None
                rd_type = self._clean(row.iloc[21]) if len(row) > 21 else None

                # Определяем тип трансмиссии из drive_trans
                transmission_type = 'MT' if 'M/T' in drive_trans else 'AT' if 'A/T' in drive_trans else 'CVT' if 'CVT' in drive_trans.upper() else None
                drive_type = '4WD' if '4WD' in drive_trans else '2WD' if '2WD' in drive_trans else None

                rec = VehicleRecord(
                    vehicle_type='passenger_car',
                    brand=brand,
                    model=model,
                    year_start=year_start,
                    year_end=year_end,
                    market='JDM',
                    attributes={
                        'displacement_liters': self._safe_displacement(displacement),
                        'engine_code': engine_code,
                        'fuel_type': 'petrol',  # JDM-каталог в основном бензин
                        'drive_type': drive_type,
                        'transmission_type': transmission_type,
                    },
                    external_codes={
                        'body_number': body_number,
                    } if body_number else {},
                    source='japan_catalog',
                )

                # Добавляем рекомендации
                if engine_oil:
                    rec.recommendations.append(FluidRecommendation(
                        node_code='ENGINE',
                        fluid_name=engine_oil,
                        fluid_brand=self._extract_brand_from_name(engine_oil),
                        viscosity_sae=engine_sae or self._parse_viscosity(engine_oil),
                        recommendation_rank=1,
                        volume_liters=engine_volume,
                        applicability_conditions={},
                    ))
                if mt_oil and 'MT' in (drive_trans or ''):
                    rec.recommendations.append(FluidRecommendation(
                        node_code='MANUAL_TRANSMISSION',
                        fluid_name=mt_oil,
                        fluid_brand=self._extract_brand_from_name(mt_oil),
                        viscosity_sae=self._parse_viscosity(mt_oil),
                        recommendation_rank=1,
                        volume_liters=mt_volume,
                    ))
                if at_oil and ('AT' in (drive_trans or '') or 'A/T' in (drive_trans or '')):
                    rec.recommendations.append(FluidRecommendation(
                        node_code='AUTO_TRANSMISSION',
                        fluid_name=at_oil,
                        fluid_brand=self._extract_brand_from_name(at_oil),
                        viscosity_sae=self._parse_viscosity(at_oil),
                        recommendation_rank=1,
                        volume_liters=at_volume,
                    ))
                if fd_type:
                    rec.recommendations.append(FluidRecommendation(
                        node_code='DIFFERENTIAL',
                        fluid_name=fd_type,
                        fluid_brand=self._extract_brand_from_name(fd_type),
                        volume_liters=fd_volume,
                        recommendation_rank=1,
                        applicability_conditions={'position': 'front'},
                    ))
                if rd_type:
                    rec.recommendations.append(FluidRecommendation(
                        node_code='DIFFERENTIAL',
                        fluid_name=rd_type,
                        fluid_brand=self._extract_brand_from_name(rd_type),
                        volume_liters=rd_volume,
                        recommendation_rank=1,
                        applicability_conditions={'position': 'rear'},
                    ))

                if rec.recommendations:
                    records.append(rec)

        logger.info(f"JapaneseCatalogParser: parsed {len(records)} vehicles from {file_path}")
        return records


# ============================================================================
# PVL Catalog Parser — для "Каталог подбора PVL.xlsx" (Газпромнефть)
# ============================================================================

class PVLCatalogParser(CatalogParser):
    """
    Парсер для каталога PVL (Газпромнефть, 31 колонка).

    Структура:
      - row 1: заголовки категорий (Легковые | Двигатель | Ручная трансмиссия | ...)
      - row 3+: бренд (AUDI, BMW, MERCEDES-BENZ, ...)
      - row 4+: модели с характеристиками
      - Diesel-секции: внутри бренда, отделены merged-row 'Diesel'
      - На каждый узел 3 колонки: основной продукт | альтернатива | объём
      - Маркеры комплектации (u, a, c, d, ...) в отдельных колонках
      - Сноски: merged rows типа 'a  Раздаточная коробка: 0,36-0,38 л G-Box GL-5 75W-90'
    """

    # Категории узлов и их колонки (по row 1)
    # [node_code, label, col_offset_main, col_offset_alt, col_offset_volume, col_offset_marker]
    NODE_COLUMNS = {
        'ENGINE':              {'main': 3,  'alt1': 4,  'alt2': 5,  'volume': 7,  'marker': 6},
        'MANUAL_TRANSMISSION': {'main': 8,  'alt1': 9,  'alt2': None, 'volume': None, 'marker': 10},
        'AUTO_TRANSMISSION':   {'main': 11, 'alt1': 12, 'alt2': None, 'volume': 14, 'marker': 13},
        'DIFFERENTIAL':        {'main': 15, 'alt1': 16, 'alt2': None, 'volume': 18, 'marker': 17},
        'COOLANT':             {'main': 19, 'alt1': 20, 'alt2': None, 'volume': 21, 'marker': None},
        'BRAKE':               {'main': 22, 'alt1': 23, 'alt2': None, 'volume': 24, 'marker': None},
        'STEERING':            {'main': 25, 'alt1': 26, 'alt2': None, 'volume': 27, 'marker': None},
        'SUSPENSION':          {'main': 28, 'alt1': 29, 'alt2': None, 'volume': 30, 'marker': None},
    }

    # Бренды, которые встречаются в файле
    KNOWN_BRANDS = {
        'AUDI', 'BMW', 'MERCEDES-BENZ', 'HONDA', 'HONDA (RUS)', 'HONDA (USA)',
        'TOYOTA', 'NISSAN', 'NISSAN (USA)', 'MITSUBISHI', 'MAZDA', 'SUBARU',
        'LEXUS', 'LEXUS (USA)', 'INFINITI', 'HYUNDAI', 'HYUNDAI (RUS)', 'HYUNDAI (USA)',
        'KIA', 'LAND ROVER', 'JAGUAR', 'JEEP', 'CHRYSLER', 'DODGE',
        'DAIHATSU', 'DAIHATSU (RUS)', 'RENAULT', 'RENAULT (RUS)',
        'CITROËN', 'PEUGEOT', 'FORD', 'FORD (RUS)', 'FORD (USA)',
        'VOLVO', 'VW', 'VOLKSWAGEN', 'OPEL', 'SKODA', 'SEAT',
        'LADA (SHIGULI/VAZ)', 'LIFAN', 'GEELY', 'CHERY', 'TAGAZ',
        'MINI', 'FIAT', 'ALFA ROMEO', 'PORSCHE', 'SAAB',
    }

    def detect(self, file_path: str) -> bool:
        try:
            df = pd.read_excel(file_path, sheet_name=0, header=None, nrows=5)
            # PVL-файл начинается с 'Легковые автомобили' в row 1
            row1 = df.iloc[1].astype(str).str.strip().tolist()
            if any('Легковые автомобили' in v for v in row1 if v != 'nan'):
                return True
            return False
        except Exception:
            return False

    def parse(self, file_path: str) -> list[VehicleRecord]:
        records: list[VehicleRecord] = []
        df = pd.read_excel(file_path, sheet_name=0, header=None)

        current_brand: Optional[str] = None
        current_model: Optional[str] = None
        current_year_start: Optional[int] = None
        current_year_end: Optional[int] = None
        current_displacement: Optional[str] = None
        in_diesel_section = False
        current_market = 'RUS'  # по умолчанию, переопределяется по бренду

        for i in range(3, len(df)):  # начинаем с row 3 (после заголовков)
            row = df.iloc[i]
            col0 = self._clean(row.iloc[0])

            if not col0:
                continue

            # Diesel-секция
            if col0 == 'Diesel':
                in_diesel_section = True
                continue

            # Бренд? (одна из строк содержит ТОЛЬКО название бренда, объединённое merged cell)
            if col0.upper().strip() in self.KNOWN_BRANDS:
                current_brand = self._normalize_brand(col0)
                # Определяем рынок
                if '(RUS)' in col0:
                    current_market = 'RU'
                elif '(USA)' in col0:
                    current_market = 'US'
                else:
                    current_market = 'EU'  # по умолчанию для PVL
                in_diesel_section = False
                continue

            # Подзаголовок поколения? (например "E46 - 3 Series")
            # Если row содержит поколение, а следующая за ним — модели
            if col0.startswith('E') and ' - ' in col0 and len(col0) < 60:
                # Это поколение BMW/Mercedes
                current_model = col0.split(' - ')[-1].strip()
                continue

            # Mercedes-Benz generations: "A-Class (W168)", "B-Class (W245)" и т.д.
            m = re.match(r'^[A-Z]+-Class\s*\([A-Z]+\d+\)', col0)
            if m and current_brand == 'Mercedes-Benz':
                current_model = col0
                continue

            # Строка-сноска? (начинается с буквы + 2 пробела)
            if re.match(r'^[a-z]{1,2}\s{2,}', col0):
                # Это сноска типа "a  Раздаточная коробка: 0,36-0,38 л G-Box GL-5 75W-90"
                # Привязываем к ТЕКУЩЕМУ auto — добавляем как альтернативную рекомендацию
                if records and col0:
                    self._apply_footnote(records[-1], col0)
                continue

            # Если текущий бренд не задан — пропускаем
            if not current_brand:
                continue

            # Это строка модели
            model_name = col0
            # Если у нас уже было поколение (например "E46 - 3 Series"), добавляем его к модели
            if current_model:
                model_name = f"{current_model} {model_name}"

            year_str = self._clean(row.iloc[1]) or ''
            year_start, year_end = self._parse_year_range(year_str)
            displacement = self._clean(row.iloc[2])

            # Создаём запись
            sub_model = 'Diesel' if in_diesel_section else None
            rec = VehicleRecord(
                vehicle_type='passenger_car',
                brand=current_brand,
                model=model_name,
                sub_model=sub_model,
                year_start=year_start,
                year_end=year_end,
                market=current_market,
                attributes={
                    'displacement_liters': self._safe_displacement(displacement),
                    'fuel_type': 'diesel' if in_diesel_section else 'petrol',
                },
                source='pvl',
            )

            # Парсим все 8 узлов
            for node_code, cols in self.NODE_COLUMNS.items():
                # Основное масло
                main_fluid = self._clean(row.iloc[cols['main']]) if cols['main'] is not None else None
                alt1_fluid = self._clean(row.iloc[cols['alt1']]) if cols['alt1'] is not None else None
                alt2_fluid = self._clean(row.iloc[cols['alt2']]) if cols['alt2'] is not None else None
                volume = self._normalize_volume(
                    self._clean(row.iloc[cols['volume']]) if cols['volume'] is not None else None
                )
                marker = self._clean(row.iloc[cols['marker']]) if cols['marker'] is not None else None

                if main_fluid or alt1_fluid or alt2_fluid:
                    # Условия применимости
                    conditions = {}
                    if marker and marker not in (' ', '·'):
                        conditions['marker_code'] = marker
                    if in_diesel_section:
                        conditions['fuel_type'] = 'diesel'

                    # Primary
                    if main_fluid:
                        rec.recommendations.append(FluidRecommendation(
                            node_code=node_code,
                            fluid_name=main_fluid,
                            fluid_brand=self._extract_brand_from_name(main_fluid),
                            viscosity_sae=self._parse_viscosity(main_fluid),
                            recommendation_rank=1,
                            is_oem=True,  # Основное — это OEM-рекомендация
                            volume_liters=volume,
                            applicability_conditions=conditions,
                        ))
                    # Альтернатива 1
                    if alt1_fluid and alt1_fluid != main_fluid:
                        rec.recommendations.append(FluidRecommendation(
                            node_code=node_code,
                            fluid_name=alt1_fluid,
                            fluid_brand=self._extract_brand_from_name(alt1_fluid),
                            viscosity_sae=self._parse_viscosity(alt1_fluid),
                            recommendation_rank=2,
                            is_oem=False,
                            volume_liters=volume,
                            applicability_conditions=conditions,
                        ))
                    # Альтернатива 2
                    if alt2_fluid and alt2_fluid != main_fluid and alt2_fluid != alt1_fluid:
                        rec.recommendations.append(FluidRecommendation(
                            node_code=node_code,
                            fluid_name=alt2_fluid,
                            fluid_brand=self._extract_brand_from_name(alt2_fluid),
                            viscosity_sae=self._parse_viscosity(alt2_fluid),
                            recommendation_rank=3,
                            is_oem=False,
                            volume_liters=volume,
                            applicability_conditions=conditions,
                        ))

            if rec.recommendations:
                records.append(rec)

        logger.info(f"PVLCatalogParser: parsed {len(records)} vehicles from {file_path}")
        return records

    def _apply_footnote(self, record: VehicleRecord, footnote: str) -> None:
        """Применяет сноску к последней записи.

        Пример: 'a  Раздаточная коробка: 0,36-0,38 л G-Box GL-5 75W-90;G-Box GL-4/GL-5 75W-90'
        """
        # Убираем маркер (1-2 буквы + 2+ пробела)
        m = re.match(r'^([a-z]{1,2})\s{2,}(.+)$', footnote)
        if not m:
            return
        marker, body = m.groups()

        # Извлекаем узел
        node_label_map = {
            'Раздаточная коробка': 'TRANSFER_CASE',
            'Раздаточная коробка G-Box': 'TRANSFER_CASE',
            'Передний дифференциал': 'DIFFERENTIAL',
            'Задний дифференциал': 'DIFFERENTIAL',
            'Самоблокирующийся дифференциал': 'DIFFERENTIAL',
            'Передний и задний дифференциалы': 'DIFFERENTIAL',
            'Дифференциал, задний (4WD)': 'DIFFERENTIAL',
            'Дифференциал, передний (4WD)': 'DIFFERENTIAL',
            'Дифференциал, модели с автоматической КПП': 'DIFFERENTIAL',
            'Вариатор (CVT)': 'CVT',
            'Автоматическая трансмиссия': 'AUTO_TRANSMISSION',
            'Коробка передач': 'MANUAL_TRANSMISSION',
            'Сцепление Haldex': 'TRANSFER_CASE',
            'PTO (4x4)': 'PTO',
            'Гидравлическая система AWC': 'STEERING',
        }
        node_code = None
        for label, code in node_label_map.items():
            if label.lower() in body.lower():
                node_code = code
                break
        if not node_code:
            return

        # Извлекаем объём
        vol_match = re.search(r'([\d,\-\s]+)\s*л', body)
        volume = self._normalize_volume(vol_match.group(1)) if vol_match else None

        # Извлекаем масла (после ';')
        oil_part = body.split(':', 1)[-1] if ':' in body else body
        oils = [o.strip() for o in oil_part.split(';') if o.strip() and o.strip() != '-']

        if not oils:
            return

        # Добавляем как alt-рекомендации
        existing_ranks = {r.recommendation_rank for r in record.recommendations
                          if r.node_code == node_code}
        next_rank = max(existing_ranks) + 1 if existing_ranks else 1

        for oil_name in oils:
            record.recommendations.append(FluidRecommendation(
                node_code=node_code,
                fluid_name=oil_name,
                fluid_brand=self._extract_brand_from_name(oil_name),
                viscosity_sae=self._parse_viscosity(oil_name),
                recommendation_rank=next_rank,
                is_oem=False,
                volume_liters=volume,
                applicability_conditions={'footnote_marker': marker, 'source': 'pvl_footnote'},
            ))
            next_rank += 1


# ============================================================================
# Future parsers — заглушки с правильным API
# ============================================================================

class HeavyEquipmentParser(CatalogParser):
    """Будущий парсер для тяжелой техники (пока не реализован)."""

    def detect(self, file_path: str) -> bool:
        # Будет определено, когда появится файл-пример
        return False

    def parse(self, file_path: str) -> list[VehicleRecord]:
        raise NotImplementedError("Heavy equipment parser — TODO when sample file available")


class SmallEngineParser(CatalogParser):
    """Будущий парсер для бензопил/газонокосилок/генераторов."""

    def detect(self, file_path: str) -> bool:
        return False

    def parse(self, file_path: str) -> list[VehicleRecord]:
        raise NotImplementedError("Small engine parser — TODO when sample file available")


# ============================================================================
# Factory — выбор подходящего парсера
# ============================================================================

class CatalogParserFactory:
    """Выбирает подходящий парсер по файлу."""

    PARSERS = [
        JapaneseCatalogParser(),
        PVLCatalogParser(),
        HeavyEquipmentParser(),
        SmallEngineParser(),
    ]

    @classmethod
    def get_parser(cls, file_path: str) -> CatalogParser:
        """Возвращает первый парсер, который detect() = True."""
        for parser in cls.PARSERS:
            try:
                if parser.detect(file_path):
                    logger.info(f"File {file_path} detected by {parser.__class__.__name__}")
                    return parser
            except Exception as e:
                logger.debug(f"{parser.__class__.__name__}.detect failed: {e}")

        raise ValueError(
            f"Unknown catalog format for file: {file_path}. "
            f"Available parsers: {[p.__class__.__name__ for p in cls.PARSERS]}"
        )

    @classmethod
    def list_supported_formats(cls) -> list[dict]:
        """Возвращает список поддерживаемых форматов для UI."""
        return [
            {
                'name': 'Japanese JDM Catalog',
                'parser': 'JapaneseCatalogParser',
                'description': 'Японские праворульные авто (Honda, Toyota, Nissan, ...)',
                'examples': ['katalog_gsm.xlsx'],
            },
            {
                'name': 'PVL Catalog (Gazpromneft)',
                'parser': 'PVLCatalogParser',
                'description': 'Каталог подбора PVL — все бренды легковых (EU, US, RU)',
                'examples': ['Каталог подбора PVL.xlsx'],
            },
            {
                'name': 'Heavy Equipment',
                'parser': 'HeavyEquipmentParser',
                'description': 'Спецтехника (экскаваторы, бульдозеры) — в разработке',
                'examples': [],
            },
            {
                'name': 'Small Engines',
                'parser': 'SmallEngineParser',
                'description': 'Бензопилы, газонокосилки, генераторы — в разработке',
                'examples': [],
            },
        ]


# ============================================================================
# Demo — запуск на текущих файлах
# ============================================================================

if __name__ == '__main__':
    import sys

    logging.basicConfig(level=logging.INFO, format='%(levelname)s %(message)s')

    files = [
        '/home/z/my-project/upload/katalog_gsm.xlsx',
        '/home/z/my-project/upload/Каталог подбора PVL.xlsx',
    ]

    for f in files:
        if not Path(f).exists():
            print(f"\n❌ File not found: {f}")
            continue

        print(f"\n{'=' * 70}")
        print(f"FILE: {f}")
        print('=' * 70)

        try:
            parser = CatalogParserFactory.get_parser(f)
            print(f"Detected parser: {parser.__class__.__name__}")
            records = parser.parse(f)
            print(f"Parsed {len(records)} vehicles")

            # Статистика
            brands = set(r.brand for r in records)
            total_recs = sum(len(r.recommendations) for r in records)
            print(f"Brands: {len(brands)} — {sorted(brands)[:10]}")
            print(f"Total recommendations: {total_recs}")

            # Show first 3 records
            print(f"\nSample records (first 3):")
            for r in records[:3]:
                print(f"  [{r.vehicle_type}] {r.brand} {r.model} ({r.year_start}-{r.year_end})")
                print(f"    Market: {r.market}, Sub: {r.sub_model}")
                print(f"    Source: {r.source}, Hash: {r.source_hash()}")
                print(f"    Attributes: {r.attributes}")
                print(f"    Recommendations ({len(r.recommendations)}):")
                for rec in r.recommendations[:5]:
                    print(f"      [{rec.node_code}] rank={rec.recommendation_rank} {rec.fluid_name} ({rec.viscosity_sae}) vol={rec.volume_liters}")
                print()

        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
